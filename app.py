"""
app.py  —  Flask web UI for Weekend Mail Automation
Place this file in the PROJECT ROOT (same level as run.py and scripts/).
Run:  python3 app.py
Then open:  http://localhost:5050
"""

from flask import Flask, render_template, jsonify, request, send_file
from werkzeug.utils import secure_filename
from pathlib import Path
import threading
import json
import builtins
import re
import time
import shutil

# ── project imports (unchanged) ──────────────────────────────────────────────
import scripts.read_mails as mail_reader
import scripts.email_report as email_report

BASE_DIR     = Path(__file__).resolve().parent
MAPPING_FILE = BASE_DIR / "customer_vertical_mapping.xlsx"
DEVICE_FILE  = BASE_DIR / "data" / "custom_devices.json"
DISABLED_DEVICE_FILE = BASE_DIR / "data" / "disabled_devices.json"
OUTPUT_DIR   = BASE_DIR / "output"
ACTIVITY_LOG_DIR   = BASE_DIR / "logs"
ACTIVITY_LOG_FILE  = ACTIVITY_LOG_DIR / "activity.log"        # always today's log — single file

EMAIL_RE     = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

app = Flask(__name__)

# ── in-memory log (thread-safe append) ───────────────────────────────────────
from collections import deque
import datetime

def _load_activity_log_if_today():
    """Restore today's log from disk if the file's first line (a date
    stamp) matches today; otherwise discard it."""
    try:
        if ACTIVITY_LOG_FILE.exists():
            lines = ACTIVITY_LOG_FILE.read_text(encoding="utf-8").splitlines()
            if lines and lines[0] == datetime.date.today().isoformat():
                return lines[1:]   # drop the date-stamp line itself
    except Exception:
        pass
    _clear_activity_log_files()
    return []

def _clear_activity_log_files():
    """Remove any on-disk log so the logs/ folder never accumulates."""
    try:
        ACTIVITY_LOG_FILE.unlink(missing_ok=True)
    except Exception:
        pass

_activity_log_date = datetime.date.today()
log_buffer = deque(_load_activity_log_if_today())
log_lock   = threading.Lock()
_scan_done_flag = False   # set to True when a scan cycle finishes; cleared by frontend poll

def _rollover_activity_log_if_needed():
    """At local midnight: clear the UI history AND delete today's file
    from disk, so a new file starts fresh and nothing piles up."""
    global _activity_log_date
    today = datetime.date.today()
    if today != _activity_log_date:
        _activity_log_date = today
        log_buffer.clear()
        _clear_activity_log_files()

def push_log(msg: str):
    ts   = datetime.datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}]  {msg}"
    with log_lock:
        _rollover_activity_log_if_needed()
        log_buffer.append(line)
        # Persist to a single file for today only, so a server restart
        # doesn't lose the day's activity. Overwritten fresh each new day.
        try:
            ACTIVITY_LOG_DIR.mkdir(exist_ok=True)
            is_new_file = not ACTIVITY_LOG_FILE.exists()
            with ACTIVITY_LOG_FILE.open("a", encoding="utf-8") as activity_file:
                if is_new_file:
                    activity_file.write(_activity_log_date.isoformat() + "\n")
                activity_file.write(line + "\n")
        except Exception:
            # Logging must never interrupt the mail automation itself.
            pass

automation_thread = None
_live_page        = None   # set by read_mails when browser opens

# ── patch read_mails to also push to log_buffer ───────────────────────────────
_orig_print = builtins.print

def _log_print(*args, **kwargs):
    global _scan_done_flag
    msg = " ".join(str(a) for a in args)
    push_log(msg)
    # Detect scan completion to trigger frontend notification
    if "scan done:" in msg.lower() or "scan complete" in msg.lower():
        _scan_done_flag = True
    _orig_print(*args, **kwargs)

mail_reader.print  = _log_print
email_report.print = _log_print


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS  (all original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

def load_custom_devices() -> dict:
    DEVICE_FILE.parent.mkdir(exist_ok=True)
    if not DEVICE_FILE.exists():
        return {}
    try:
        return json.loads(DEVICE_FILE.read_text())
    except Exception:
        return {}

def save_custom_devices(data: dict):
    DEVICE_FILE.parent.mkdir(exist_ok=True)
    DEVICE_FILE.write_text(json.dumps(data, indent=2))

def load_disabled_devices() -> set:
    if not DISABLED_DEVICE_FILE.exists():
        return set()
    try:
        return {
            str(keyword).strip().lower()
            for keyword in json.loads(DISABLED_DEVICE_FILE.read_text())
            if str(keyword).strip()
        }
    except Exception:
        return set()

def save_disabled_devices(keywords: set):
    DISABLED_DEVICE_FILE.parent.mkdir(exist_ok=True)
    DISABLED_DEVICE_FILE.write_text(json.dumps(sorted(keywords), indent=2))

def load_customers() -> list:
    if not MAPPING_FILE.exists():
        return []

    try:
        import pandas as pd

        df = pd.read_excel(MAPPING_FILE, engine="openpyxl")
        df.columns = [str(c).strip().lower() for c in df.columns]

        cc = next((c for c in df.columns if "customer" in c or "company" in c), None)
        vc = next((c for c in df.columns if "vertical" in c), None)
        mc = next((c for c in df.columns if "manager" in c or "responsible" in c), None)

        if not cc or not vc:
            return []

        return [
            {
                "customer": str(r[cc]).strip(),
                "vertical": str(r[vc]).strip(),
                "manager": (
                    str(r[mc]).strip()
                    if mc and str(r[mc]).strip().lower() != "nan"
                    else ""
                ),
            }
            for _, r in df.iterrows()
            if str(r[cc]).strip() not in ("", "nan")
        ]

    except Exception:
        return []

def append_customer(customer: str, vertical: str, manager: str = "") -> bool:
    try:
        import pandas as pd
        from openpyxl import load_workbook
        if not MAPPING_FILE.exists():
            pd.DataFrame(columns=["Customer", "Vertical"]).to_excel(
                MAPPING_FILE, index=False, engine="openpyxl")
        wb = load_workbook(MAPPING_FILE)
        wb.active.append([vertical, customer, manager])
        wb.save(MAPPING_FILE)
        try:
            from scripts.vertical_lookup import load_vertical_mapping
            load_vertical_mapping()
        except Exception:
            pass
        return True
    except Exception as e:
        return False

TRACKER_FILE = BASE_DIR / "output" / "Weekend_Cases_Tracker.xlsx"

def latest_excel():
    return TRACKER_FILE if TRACKER_FILE.exists() else None

def email_status() -> dict:
    try:
        recipients = load_email_recipients()
        return {
            "configured": bool(recipients),
            "recipients": recipients,
            "message": (
                f"{len(recipients)} recipient(s) configured"
                if recipients else
                "No recipients configured"
            )
        }
    except FileNotFoundError:
        return {
            "configured": False,
            "recipients": [],
            "message": "email_config.json not found"
        }
    except Exception as e:
        return {
            "configured": False,
            "recipients": [],
            "message": f"Email config error: {e}"
        }

def load_email_config() -> dict:
    if not email_report.CONFIG_FILE.exists():
        return {"sender_email": "", "recipients": []}
    try:
        data = json.loads(email_report.CONFIG_FILE.read_text())
    except Exception:
        data = {}
    data.setdefault("sender_email", "")
    data.setdefault("recipients", [])
    return data

def save_email_config(data: dict):
    email_report.CONFIG_FILE.write_text(json.dumps(data, indent=4))

def load_email_recipients() -> list:
    config = load_email_config()
    disabled = {
        str(email).strip().lower()
        for email in config.get("disabled_recipients", [])
    }
    seen = set()
    recipients = []
    for email in config.get("recipients", []):
        email = str(email).strip()
        key = email.lower()
        if email and key not in seen and key not in disabled:
            recipients.append(email)
            seen.add(key)
    return recipients

def load_all_email_recipients() -> list:
    config = load_email_config()
    disabled = {
        str(email).strip().lower()
        for email in config.get("disabled_recipients", [])
    }
    seen = set()
    recipients = []
    for email in config.get("recipients", []):
        email = str(email).strip()
        key = email.lower()
        if email and key not in seen:
            recipients.append({"email": email, "enabled": key not in disabled})
            seen.add(key)
    return recipients

BUILTIN_DEVICES = {
    "mx":"Routing","ptx":"Routing","acx":"Routing",
    "srx":"Security","ssg":"Security",
    "qfx":"Switching","ex":"Switching",
    "mist":"Wireless","software":"Software",
}


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — PAGES  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    xl = latest_excel()
    return render_template("index.html",
                           latest_file=xl.name if xl else "—",
                           running=mail_reader.RUNNING)


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — AUTOMATION API  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/start", methods=["POST"])
def api_start():
    global automation_thread

    if automation_thread and not automation_thread.is_alive():
        automation_thread = None

    if automation_thread and automation_thread.is_alive():
        return jsonify({"ok": False, "msg": "Already running"})

    data = request.get_json(silent=True) or {}

    def folder_list(value):
        values = value if isinstance(value, list) else [value]
        folders = [str(item).strip() for item in values if str(item).strip()]
        return folders or ["inbox"]

    dispatch_folders = folder_list(
        data.get("dispatch_folders", data.get("dispatch_folder", "inbox"))
    )
    handover_folders = folder_list(
        data.get("handover_folders", data.get("handover_folder", "inbox"))
    )
    em_name = str(data.get("em_name", "")).strip()

    mail_reader.RUNNING = True
    push_log(
        f"Automation started. Dispatch: {dispatch_folders} | "
        f"Handover: {handover_folders} | EM: '{em_name}'"
    )

    automation_thread = threading.Thread(
        target=mail_reader.run_mail_reader,
        kwargs={
            "dispatch_folders": dispatch_folders,
            "handover_folders": handover_folders,
            "em_name":          em_name,
        },
        daemon=True
    )
    automation_thread.start()
    return jsonify({"ok": True})

@app.route("/api/stop", methods=["POST"])
def api_stop():
    global automation_thread

    # If a report is currently being sent (manual OR automatic — both set
    # _paused now), do NOT force-close the browser. Killing it mid-send
    # leaves a half-filled compose window / partially-typed recipients and
    # corrupts the send, which is what was forcing a manual restart.
    # Wait for the send to finish first; poll every second, up to 2 minutes.
    waited = 0
    while getattr(mail_reader, "_paused", False) and waited < 120:
        push_log("Stop requested — report is still sending, waiting for it to finish...") if waited == 0 else None
        time.sleep(1)
        waited += 1

    if getattr(mail_reader, "_paused", False):
        # Still sending after 2 minutes — something's stuck. Refuse to
        # force-stop rather than risk corrupting/duplicating the send;
        # the caller can retry once it clears.
        return jsonify({
            "ok": False,
            "msg": "Report is still being sent after waiting 2 minutes — "
                   "please wait for it to finish before stopping."
        }), 409

    mail_reader.RUNNING = False

    # Force-close the browser so the thread exits immediately
    try:
        lp = getattr(mail_reader, "_live_page", None)
        if lp is not None:
            lp.context.browser.close()
    except Exception:
        pass

    # Wait up to 5 seconds for thread to finish
    if automation_thread and automation_thread.is_alive():
        automation_thread.join(timeout=5)

    automation_thread    = None
    mail_reader.RUNNING  = False
    push_log("Automation stopped by user.")
    return jsonify({"ok": True})

@app.route("/api/status")
def api_status():
    global _scan_done_flag
    alive = bool(automation_thread and automation_thread.is_alive())
    paused = getattr(mail_reader, "_paused", False)
    email_ready = alive and getattr(mail_reader, "_live_page", None) is not None
    if not alive and mail_reader.RUNNING:
        mail_reader.RUNNING = False
    with log_lock:
        _rollover_activity_log_if_needed()
        log_count = len(log_buffer)
        try:
            requested_start = int(request.args.get("logs_since", 0))
            log_start = requested_start if 0 <= requested_start <= log_count else 0
        except (TypeError, ValueError):
            log_start = 0
        logs = list(log_buffer)[log_start:]
    xl = latest_excel()
    sleeping = getattr(mail_reader, "_sleeping", False)
    # Consume scan_done flag (one-shot — cleared after frontend reads it)
    scan_done = _scan_done_flag
    if _scan_done_flag:
        _scan_done_flag = False
    return jsonify({
        "running":       mail_reader.RUNNING,
        "paused":        paused,
        "sleeping":      sleeping,
        "logs":          logs,
        "log_start":     log_start,
        "log_count":     log_count,
        "latest_file":   xl.name if xl else "—",
        "email":         email_status(),
        "email_sending": paused,
        "email_ready":   sleeping and not paused,
        "scan_done":     scan_done,
    })

@app.route("/api/send_report", methods=["POST"])
def api_send_report():
    if not mail_reader.RUNNING:
        return jsonify({"ok": False, "msg": "Automation is not running."}), 400

    if not getattr(mail_reader, "_sleeping", False):
        return jsonify({"ok": False, "msg": "Scan is in progress — wait for the 1-min sleep window, then try again."}), 400

    page = getattr(mail_reader, "_live_page", None)
    if page is None:
        return jsonify({"ok": False, "msg": "Browser not ready."}), 400

    # Set flag and wake the sleep so report sends immediately
    mail_reader._send_requested = True
    wake = getattr(mail_reader, "_wake_event", None)
    if wake:
        wake.set()

    push_log("Manual report requested — will send during current sleep window...")
    return jsonify({"ok": True, "msg": "Sending now — automation will resume after"})


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — EMAIL RECIPIENTS API  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/email_recipients")
def api_email_recipients():
    return jsonify(load_email_recipients())

@app.route("/api/email_recipients/details")
def api_email_recipient_details():
    return jsonify(load_all_email_recipients())

@app.route("/api/email_recipients", methods=["POST"])
def api_add_email_recipient():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip()

    if not email:
        return jsonify({"ok": False, "msg": "Email address is required"}), 400

    if not EMAIL_RE.match(email):
        return jsonify({"ok": False, "msg": "Enter a valid email address"}), 400

    config = load_email_config()
    recipients = [item["email"] for item in load_all_email_recipients()]

    if email.lower() in {r.lower() for r in recipients}:
        return jsonify({"ok": False, "msg": "Recipient already exists"}), 400

    recipients.append(email)
    config["recipients"] = recipients
    config["disabled_recipients"] = [
        item for item in config.get("disabled_recipients", [])
        if str(item).strip().lower() != email.lower()
    ]
    save_email_config(config)
    push_log(f"Email recipient added: {email}")
    return jsonify({"ok": True})

@app.route("/api/email_recipients/remove", methods=["POST"])
def api_remove_email_recipient():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip()

    recipients = [item["email"] for item in load_all_email_recipients()]
    kept = [r for r in recipients if r.lower() != email.lower()]

    if len(kept) == len(recipients):
        return jsonify({"ok": False, "msg": "Recipient not found"}), 404

    config = load_email_config()
    config["recipients"] = kept
    config["disabled_recipients"] = [
        item for item in config.get("disabled_recipients", [])
        if str(item).strip().lower() != email.lower()
    ]
    save_email_config(config)
    push_log(f"Email recipient removed: {email}")
    return jsonify({"ok": True})

@app.route("/api/email_recipients/toggle", methods=["POST"])
def api_toggle_email_recipient():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip()
    enabled = bool(data.get("enabled"))
    recipients = [item["email"] for item in load_all_email_recipients()]

    if email.lower() not in {item.lower() for item in recipients}:
        return jsonify({"ok": False, "msg": "Recipient not found"}), 404

    config = load_email_config()
    disabled = {
        str(item).strip().lower()
        for item in config.get("disabled_recipients", [])
        if str(item).strip()
    }
    if enabled:
        disabled.discard(email.lower())
    else:
        disabled.add(email.lower())
    config["disabled_recipients"] = sorted(disabled)
    save_email_config(config)
    push_log(f"Email recipient {'enabled' if enabled else 'disabled'}: {email}")
    return jsonify({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — EXCEL  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/download_excel")
def download_excel():
    xl = latest_excel()
    if not xl:
        return jsonify({"error": "No Excel file found"}), 404
    return send_file(xl, as_attachment=True, download_name=xl.name)


UPLOAD_REQUIRED_COLUMNS = {
    "Date", "Case#", "Customer", "Vertical", "Technology", "Case Delivery Type"
}


@app.route("/api/upload_excel", methods=["POST"])
def api_upload_excel():
    """
    Lets the user upload a manually-corrected copy of the tracker
    (after fixing red/amber-highlighted rows) and makes it the new
    live tracker file. Highlighting + the Charts tab are rebuilt
    against the corrected data so future scans, charts, and email
    reports all build on the fixed version instead of the original
    mistakes.
    """
    if "file" not in request.files or not request.files["file"].filename:
        return jsonify({"ok": False, "msg": "No file selected"}), 400

    upload = request.files["file"]
    filename = secure_filename(upload.filename)
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "msg": "Please upload a .xlsx file"}), 400

    OUTPUT_DIR.mkdir(exist_ok=True)
    tmp_path = OUTPUT_DIR / f".upload_tmp_{filename}"
    try:
        upload.save(tmp_path)
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Could not save uploaded file: {e}"}), 400

    # ── Validate structure before touching the live tracker file ──────────
    try:
        import pandas as pd
        xl = pd.ExcelFile(tmp_path, engine="openpyxl")
        data_sheets = [s for s in xl.sheet_names if s in ("Sat", "Sun")]
        if not data_sheets:
            raise ValueError("File must contain a 'Sat' and/or 'Sun' sheet")

        for sheet in data_sheets:
            df = pd.read_excel(tmp_path, sheet_name=sheet, engine="openpyxl", nrows=0)
            missing = UPLOAD_REQUIRED_COLUMNS - {str(c).strip() for c in df.columns}
            if missing:
                raise ValueError(
                    f"Sheet '{sheet}' is missing required column(s): {', '.join(sorted(missing))}"
                )
    except Exception as e:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        return jsonify({"ok": False, "msg": f"Upload rejected — {e}"}), 400

    # ── Overwrite the live tracker with the corrected file ────────────────
    shutil.move(str(tmp_path), str(TRACKER_FILE))

    # ── Re-apply red/amber highlighting and rebuild Charts tab ────────────
    try:
        from openpyxl import load_workbook
        from scripts.excel_writer import _highlight_problem_rows, _rebuild_charts_sheet
        wb = load_workbook(TRACKER_FILE)
        _highlight_problem_rows(wb)
        _rebuild_charts_sheet(wb)
        wb.save(TRACKER_FILE)
    except Exception as e:
        push_log(f"Corrected Excel uploaded, but highlight/chart refresh failed: {e}")

    push_log("Corrected Excel uploaded by user — tracker file replaced.")
    return jsonify({"ok": True, "msg": "Tracker file updated."})


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — CUSTOMERS API  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/customers")
def api_customers():
    return jsonify(load_customers())

@app.route("/api/customers", methods=["POST"])
def api_add_customer():
    data     = request.get_json()
    customer = (data.get("customer") or "").strip()
    vertical = (data.get("vertical") or "").strip()
    manager  = (data.get("manager") or "").strip()

    if not customer or not vertical:
        return jsonify({"ok": False, "msg": "Customer and vertical are required"}), 400

    ok = append_customer(customer, vertical, manager)

    if ok:
        push_log(f"Customer added: {customer} → {vertical} ({manager})")

    return jsonify({"ok": ok})


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — DEVICES API  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/devices")
def api_devices():
    disabled = load_disabled_devices()
    builtin = [{"keyword": k, "technology": v, "source": "built-in"}
               for k, v in BUILTIN_DEVICES.items() if k not in disabled]
    custom  = [{"keyword": k, "technology": v, "source": "custom"}
               for k, v in load_custom_devices().items()]
    return jsonify(builtin + custom)

@app.route("/api/devices", methods=["POST"])
def api_add_device():
    data    = request.get_json()
    keyword = (data.get("keyword") or "").strip().lower()
    tech    = (data.get("technology") or "").strip()
    if not keyword or not tech:
        return jsonify({"ok": False, "msg": "Keyword and technology are required"}), 400
    if keyword in BUILTIN_DEVICES:
        return jsonify({"ok": False, "msg": f"'{keyword}' is a built-in device"}), 400
    customs = load_custom_devices()
    customs[keyword] = tech
    save_custom_devices(customs)
    try:
        from scripts import parser as p
        p.DEVICE_TECH_MAP[keyword] = tech
    except Exception:
        pass
    push_log(f"Device added: {keyword} → {tech}")
    return jsonify({"ok": True})

@app.route("/api/devices/remove", methods=["POST"])
def api_remove_device():
    data = request.get_json() or {}
    keyword = (data.get("keyword") or "").strip().lower()
    customs = load_custom_devices()

    if keyword in customs:
        del customs[keyword]
        save_custom_devices(customs)
    elif keyword in BUILTIN_DEVICES:
        disabled = load_disabled_devices()
        disabled.add(keyword)
        save_disabled_devices(disabled)
    else:
        return jsonify({"ok": False, "msg": "Device not found"}), 404
    try:
        from scripts import parser as p
        p.DEVICE_TECH_MAP.pop(keyword, None)
    except Exception:
        pass
    push_log(f"Device removed: {keyword}")
    return jsonify({"ok": True})


# ═════════════════════════════════════════════════════════════════════════════
# ── NEW: CHART DATA API ───────────────────────────────────────────────────
# Reads the latest Excel output file and returns count breakdowns.
# Does NOT touch read_mails.py, excel_writer.py, or any existing logic.
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/chart_data")
def api_chart_data():
    import pandas as pd
    import datetime as _dt

    xl = latest_excel()
    if not xl:
        return jsonify({
            "file": None, "total": 0,
            "latest":  {"vertical": {}, "technology": {}, "delivery_type": {}},
            "weekly":  {"labels": ["Sat","Sun"], "series": {}, "totals": []},
            "monthly": {"labels": ["Week 1","Week 2","Week 3","Week 4","Week 5"], "series": {}, "totals": []}
        })

    try:
        DEFAULT_VERTICAL   = ["EMEA","Cable","Content","Enterprise","Telco","Software"]
        DEFAULT_TECHNOLOGY = ["Routing","Switching","Security","Software"]
        DEFAULT_DELIVERY   = ["Dispatch P1","Dispatch P2","Handover"]

        # ── Mac-safe date parser ─────────────────────────────────────────────
        # Handles: datetime objects, '2-May-26', '02-May-26', '2-May-2026', ISO
        # Does NOT use %-d (Linux-only). Zero-pads the day before strptime.
        def _parse_one(v):
            if v is None:
                return pd.NaT
            try:
                if pd.isna(v):
                    return pd.NaT
            except Exception:
                pass
            if isinstance(v, (_dt.datetime, _dt.date)):
                return pd.Timestamp(v)
            s = str(v).strip()
            if not s or s.lower() in ("nat","none","nan",""):
                return pd.NaT
            # Zero-pad the day part so '2-May-26' → '02-May-26'
            # Handles both '-' and '/' separators
            import re
            s_padded = re.sub(
                r'^(\d{1})([-/])([A-Za-z]+)([-/])(\d{2,4})$',
                lambda m: f"0{m.group(1)}{m.group(2)}{m.group(3)}{m.group(4)}{m.group(5)}",
                s
            )
            for candidate in (s_padded, s):
                for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        return pd.Timestamp(_dt.datetime.strptime(candidate, fmt))
                    except ValueError:
                        pass
            # Last resort — pandas inference
            try:
                return pd.Timestamp(pd.to_datetime(s, dayfirst=True, errors="coerce"))
            except Exception:
                return pd.NaT

        def _parse_col(series):
            if pd.api.types.is_datetime64_any_dtype(series):
                return pd.to_datetime(series, utc=False, errors="coerce").dt.tz_localize(None)
            return pd.Series([_parse_one(v) for v in series], index=series.index, dtype="datetime64[ns]")

        # ── Value normalisation helper — defined early, used at read time ────
        ACRONYMS = {"Emea": "EMEA", "Cfts": "CFTS", "Bngl": "BNGL"}
        ALIASES_LOWER = {
            "entfin":            "Enterprise",
            "cloud":             "Software",
            "new dispatch p1":   "Dispatch P1",
            "new dispatch p2":   "Dispatch P2",
            "handover in":       "Handover",
            "handover-in":       "Handover",
        }

        def fix_case(val):
            raw = str(val).replace('\xa0', '').strip()
            if not raw or raw.lower() in ("nan", "none", "nat"):
                return ""
            normalized = " ".join(raw.lower().split())
            alias = ALIASES_LOWER.get(normalized)
            if alias:
                return alias
            if normalized.startswith("new ") and "dispatch" in normalized:
                return normalized[4:].strip().title()
            v = raw.title()
            return ACRONYMS.get(v, v)

        # ── Read both sheets ─────────────────────────────────────────────────
        dfs = []
        sheet_errors = []
        for sheet_name in ("Sat", "Sun"):
            try:
                df = pd.read_excel(xl, sheet_name=sheet_name, engine="openpyxl")
                if df.empty:
                    continue
                df["_sheet"] = sheet_name
                df["Date"]   = _parse_col(df["Date"]) if "Date" in df.columns else pd.NaT
                # Normalise categories right at read time
                for col in ("Vertical", "Technology", "Case Delivery Type"):
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.strip().apply(fix_case)
                dfs.append(df)
            except Exception as e:
                sheet_errors.append(f"{sheet_name}: {e}")

        if not dfs:
            return jsonify({
                "file": xl.name, "total": 0, "debug_errors": sheet_errors,
                "latest":  {"vertical": {}, "technology": {}, "delivery_type": {}},
                "weekly":  {"labels": ["Sat","Sun"], "series": {}, "totals": []},
                "monthly": {"labels": ["Week 1","Week 2","Week 3","Week 4","Week 5"], "series": {}, "totals": []}
            })

        combined = pd.concat(dfs, ignore_index=True)
        combined = combined[combined["Date"].notna()].copy()
        # Keep ONLY Saturday (5) and Sunday (6) — ignore weekday test data
        combined = combined[combined["Date"].dt.weekday.isin([5, 6])].copy()

        if combined.empty:
            return jsonify({
                "file": xl.name, "total": 0, "debug": "all dates parsed as NaT",
                "latest":  {"vertical": {}, "technology": {}, "delivery_type": {}},
                "weekly":  {"labels": ["Sat","Sun"], "series": {}, "totals": []},
                "monthly": {"labels": ["Week 1","Week 2","Week 3","Week 4","Week 5"], "series": {}, "totals": []}
            })

        # ── Derived columns ──────────────────────────────────────────────────
        combined["weekday"]       = combined["Date"].dt.weekday
        combined["day_name"]      = combined["weekday"].map({5: "Sat", 6: "Sun"}).fillna(
                                        combined["Date"].dt.strftime("%a"))
        combined["weekend_start"] = combined["Date"] - pd.to_timedelta(
                                        (combined["weekday"] - 5) % 7, unit="d")
        combined["week_of_month"] = (combined["Date"].dt.day - 1) // 7 + 1
        combined["week_label"]    = "Week " + combined["week_of_month"].astype(str)

        latest_weekend = combined["weekend_start"].max()
        weekly_df      = combined[combined["weekend_start"] == latest_weekend].copy()
        latest_df      = weekly_df

        # ── Filter monthly to CURRENT month only ─────────────────────────────
        now           = _dt.datetime.now()
        current_month = now.month
        current_year  = now.year
        monthly_df    = combined[
            (combined["Date"].dt.month == current_month) &
            (combined["Date"].dt.year  == current_year)
        ].copy()

        # ── Helpers ──────────────────────────────────────────────────────────
        def normalize_counts(df, col, defaults=None):
            if col not in df.columns or df.empty:
                return {k: 0 for k in (defaults or [])}
            s = df[col].dropna().astype(str).str.strip()
            s = s[s != ""]
            counts = s.value_counts().to_dict()
            result = {}
            for k in (defaults or []):
                result[str(k)] = int(counts.pop(k, 0))
            for k, v in sorted(counts.items()):
                result[str(k)] = int(v)
            return result

        def build_timeseries(df, group_col, category_col, defaults=None, labels=None):
            empty_series = {k: [0]*len(labels or []) for k in (defaults or [])}
            if category_col not in df.columns or df.empty:
                return {"labels": labels or [], "series": empty_series, "totals": [0]*len(labels or [])}
            df2 = df[[group_col, category_col]].copy()
            df2 = df2[df2[category_col].astype(str).str.strip() != ""]
            grouped = df2.groupby([group_col, category_col]).size().unstack(fill_value=0)
            categories = list(defaults) if defaults else []
            for cat in grouped.columns.astype(str):
                if cat not in categories:
                    categories.append(cat)
            grouped = grouped.reindex(columns=categories, fill_value=0)
            if labels is not None:
                grouped = grouped.reindex(index=labels, fill_value=0)
            else:
                grouped = grouped.sort_index()
            return {
                "labels": [str(l) for l in grouped.index],
                "series": {str(c): [int(grouped.at[l, c]) for l in grouped.index] for c in categories},
                "totals": [int(x) for x in grouped.sum(axis=1)]
            }

        weekly_labels  = ["Sat", "Sun"]
        monthly_labels = ["Week 1", "Week 2", "Week 3", "Week 4", "Week 5"]

        # Week date ranges for tooltip — Mac-safe (no %-d)
        week_date_ranges = {}
        for wk in range(1, 6):
            wk_df = monthly_df[monthly_df["week_of_month"] == wk]
            if not wk_df.empty:
                min_d = wk_df["Date"].min()
                max_d = wk_df["Date"].max()
                min_s = f"{min_d.day} {min_d.strftime('%b')}"
                max_s = f"{max_d.day} {max_d.strftime('%b')}"
                week_date_ranges[f"Week {wk}"] = f"Week {wk} ({min_s}–{max_s})"
            else:
                week_date_ranges[f"Week {wk}"] = f"Week {wk}"

        sat_date = weekly_df[weekly_df["day_name"] == "Sat"]["Date"].min()
        sun_date = weekly_df[weekly_df["day_name"] == "Sun"]["Date"].min()
        weekly_date_ranges = {
            "Sat": f"Sat ({sat_date.day} {sat_date.strftime('%b')})" if pd.notna(sat_date) else "Sat",
            "Sun": f"Sun ({sun_date.day} {sun_date.strftime('%b')})" if pd.notna(sun_date) else "Sun",
        }

        return jsonify({
            "file":  xl.name,
            "total": len(latest_df),
            "latest": {
                "vertical":      normalize_counts(latest_df, "Vertical",           DEFAULT_VERTICAL),
                "technology":    normalize_counts(latest_df, "Technology",         DEFAULT_TECHNOLOGY),
                "delivery_type": normalize_counts(latest_df, "Case Delivery Type", DEFAULT_DELIVERY),
            },
            "weekly": {
                "vertical":      build_timeseries(weekly_df,  "day_name",   "Vertical",           DEFAULT_VERTICAL,   weekly_labels),
                "technology":    build_timeseries(weekly_df,  "day_name",   "Technology",         DEFAULT_TECHNOLOGY, weekly_labels),
                "delivery_type": build_timeseries(weekly_df,  "day_name",   "Case Delivery Type", DEFAULT_DELIVERY,   weekly_labels),
                "date_ranges":   weekly_date_ranges,
            },
            "monthly": {
                "vertical":      build_timeseries(monthly_df, "week_label", "Vertical",           DEFAULT_VERTICAL,   monthly_labels),
                "technology":    build_timeseries(monthly_df, "week_label", "Technology",         DEFAULT_TECHNOLOGY, monthly_labels),
                "delivery_type": build_timeseries(monthly_df, "week_label", "Case Delivery Type", DEFAULT_DELIVERY,   monthly_labels),
                "date_ranges":   week_date_ranges,
            },
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ═════════════════════════════════════════════════════════════════════════════
# ── MONTHLY COMPARE ──────────────────────────────────────────────────────────
# Reads Saturday + Sunday sheets from the persistent tracker file and
# returns per-month totals + breakdowns for Vertical, Technology, Delivery.
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/api/debug_values")
def api_debug_values():
    import pandas as pd
    xl = latest_excel()
    if not xl:
        return jsonify({"error": "No file"})
    result = {}
    for sheet in ("Sat", "Sun"):
        try:
            df = pd.read_excel(xl, sheet_name=sheet, engine="openpyxl")
            for col in ("Vertical", "Technology", "Case Delivery Type"):
                if col in df.columns:
                    vals = df[col].dropna().astype(str).unique().tolist()
                    result[f"{sheet}_{col}"] = vals
        except Exception as e:
            result[sheet] = str(e)
    return jsonify(result)



@app.route("/api/debug_excel")
def api_debug_excel():
    """Hit this in browser to see exactly what's in the Excel file."""
    import pandas as pd
    xl = latest_excel()
    if not xl:
        return jsonify({"error": "No Excel file found"})
    result = {}
    try:
        wb_sheets = pd.ExcelFile(xl, engine="openpyxl").sheet_names
        result["sheets_found"] = wb_sheets
        for sheet in wb_sheets:
            df = pd.read_excel(xl, sheet_name=sheet, engine="openpyxl")
            result[sheet] = {
                "rows": len(df),
                "columns": list(df.columns),
                "first_5_dates": [str(v) for v in df["Date"].head(5).tolist()] if "Date" in df.columns else [],
                "date_dtype": str(df["Date"].dtype) if "Date" in df.columns else "no Date col",
                "delivery_type_raw": [repr(v) for v in df["Case Delivery Type"].dropna().unique().tolist()] if "Case Delivery Type" in df.columns else [],
                "vertical_raw": [repr(v) for v in df["Vertical"].dropna().unique().tolist()] if "Vertical" in df.columns else [],
            }
    except Exception as e:
        result["error"] = str(e)
    return jsonify(result)



@app.route("/api/monthly_compare")
def api_monthly_compare():
    import pandas as pd
    import datetime as _dt
    import re

    xl = latest_excel()
    if not xl:
        return jsonify({"months": [], "totals": [], "vertical": {}, "technology": {}, "delivery_type": {}})

    try:
        def _parse_one(v):
            if v is None:
                return pd.NaT
            try:
                if pd.isna(v):
                    return pd.NaT
            except Exception:
                pass
            if isinstance(v, (_dt.datetime, _dt.date)):
                return pd.Timestamp(v)
            s = str(v).strip()
            if not s or s.lower() in ("nat","none","nan",""):
                return pd.NaT
            s_padded = re.sub(
                r'^(\d{1})([-/])([A-Za-z]+)([-/])(\d{2,4})$',
                lambda m: f"0{m.group(1)}{m.group(2)}{m.group(3)}{m.group(4)}{m.group(5)}",
                s
            )
            for candidate in (s_padded, s):
                for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        return pd.Timestamp(_dt.datetime.strptime(candidate, fmt))
                    except ValueError:
                        pass
            try:
                return pd.Timestamp(pd.to_datetime(s, dayfirst=True, errors="coerce"))
            except Exception:
                return pd.NaT

        def _parse_col(series):
            if pd.api.types.is_datetime64_any_dtype(series):
                return pd.to_datetime(series, utc=False, errors="coerce").dt.tz_localize(None)
            return pd.Series([_parse_one(v) for v in series], index=series.index, dtype="datetime64[ns]")

        dfs = []
        for sheet in ("Sat", "Sun"):
            try:
                df = pd.read_excel(xl, sheet_name=sheet, engine="openpyxl")
                if df.empty:
                    continue
                df["Date"] = _parse_col(df["Date"]) if "Date" in df.columns else pd.NaT
                dfs.append(df)
            except Exception:
                continue

        if not dfs:
            return jsonify({"months": [], "totals": [], "vertical": {}, "technology": {}, "delivery_type": {}})

        combined = pd.concat(dfs, ignore_index=True)
        combined = combined[combined["Date"].notna()].copy()
        # Keep ONLY Saturday (5) and Sunday (6)
        combined = combined[combined["Date"].dt.weekday.isin([5, 6])].copy()

        if combined.empty:
            return jsonify({"months": [], "totals": [], "vertical": {}, "technology": {}, "delivery_type": {}})

        combined["month_label"] = combined["Date"].dt.strftime("%b %Y")
        combined["month_order"] = combined["Date"].dt.to_period("M")

        month_order = (
            combined[["month_label","month_order"]].drop_duplicates()
            .sort_values("month_order")["month_label"].tolist()
        )
        totals = [int((combined["month_label"] == m).sum()) for m in month_order]

        ALIASES_LOWER = {
            "entfin":            "Enterprise",
            "cloud":             "Software",
            "new dispatch p1":   "Dispatch P1",
            "new dispatch p2":   "Dispatch P2",
            "new dispatch  p1":  "Dispatch P1",
            "new dispatch  p2":  "Dispatch P2",
            "handover in":       "Handover",
            "handover-in":       "Handover",
        }

        def _fix(val):
            raw = str(val).strip()
            normalized = " ".join(raw.lower().split())
            alias = ALIASES_LOWER.get(normalized)
            if alias:
                return alias
            if normalized.startswith("new ") and "dispatch" in normalized:
                return normalized[4:].strip().title()
            v = raw.title()
            return {"Emea": "EMEA"}.get(v, v)

        def category_by_month(col):
            if col not in combined.columns:
                return {}
            df2 = combined.copy()
            df2[col] = df2[col].astype(str).str.strip().apply(_fix)
            df2 = df2[df2[col] != ""]
            pivot = (
                df2.groupby(["month_label", col]).size()
                .unstack(fill_value=0)
                .reindex(index=month_order, fill_value=0)
            )
            return {str(c): [int(pivot.at[m,c]) for m in month_order] for c in pivot.columns.astype(str)}

        return jsonify({
            "months":        month_order,
            "totals":        totals,
            "vertical":      category_by_month("Vertical"),
            "technology":    category_by_month("Technology"),
            "delivery_type": category_by_month("Case Delivery Type"),
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ═════════════════════════════════════════════════════════════════════════════
# STARTUP — load custom devices into parser  (original — untouched)
# ═════════════════════════════════════════════════════════════════════════════

def _startup():
    try:
        from scripts import parser as p
        for kw in load_disabled_devices():
            p.DEVICE_TECH_MAP.pop(kw, None)
        for kw, tech in load_custom_devices().items():
            p.DEVICE_TECH_MAP[kw] = tech
    except Exception:
        pass
    push_log("Web UI ready. Open http://localhost:5050 in your browser.")

_startup()
 

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=False, use_reloader=False, threaded=True)