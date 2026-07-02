import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

BASE_DIR    = Path(__file__).resolve().parent.parent
OUTPUT_FILE = BASE_DIR / "output" / "Weekend_Cases_Tracker.xlsx"

SHEET_SATURDAY = "Sat"
SHEET_SUNDAY   = "Sun"

COLUMNS = [
    "Date", "Case#", "Customer", "Vertical",
    "Technology", "Case Delivery Type", "EM", "Comments"
]

CHART_COLORS = [
    "4472C4", "ED7D31", "A9D18E",
    "FF0000", "FFC000", "70AD47",
    "9E480E", "997300"
]

CRITICAL_COLS = [2, 3, 4, 5, 6]   # Case#, Customer, Vertical, Technology, Delivery
MISSING_FILL = PatternFill("solid", start_color="FFB3B3")
MISSING_FONT = Font(name="Arial", color="7B0000")
DUPLICATE_FILL = PatternFill("solid", start_color="FFD966")
DUPLICATE_FONT = Font(name="Arial", color="7F6000")
PROBLEM_FILL_COLORS = {
    "00FFB3B3", "FFFFB3B3",
    "00FFD966", "FFFFD966",
}


def get_sheet_name(dt=None):
    weekday = (dt or datetime.now()).weekday()
    if weekday == 5:
        return SHEET_SATURDAY
    elif weekday == 6:
        return SHEET_SUNDAY
    else:
        return SHEET_SATURDAY  # TEST_MODE / weekday run


def get_output_file():
    return OUTPUT_FILE


def _ensure_workbook():
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    if not OUTPUT_FILE.exists():
        wb = Workbook()
        ws_sat = wb.active
        ws_sat.title = SHEET_SATURDAY
        ws_sat.append(COLUMNS)
        ws_sun = wb.create_sheet(SHEET_SUNDAY)
        ws_sun.append(COLUMNS)
        wb.save(OUTPUT_FILE)
        return wb
    wb = load_workbook(OUTPUT_FILE)
    changed = False
    for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(COLUMNS)
            changed = True
    if changed:
        wb.save(OUTPUT_FILE)
    return wb


def _adjust_column_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        col_idx    = column_cells[0].column
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 5, 50)


def _rebuild_charts_sheet(wb):
    # ── Same normalization as chart_exporter.py ──────────────────────────
    _ACRONYMS = {"Emea": "EMEA", "Cfts": "CFTS", "Bngl": "BNGL"}
    _ALIASES  = {
        "entfin":            "Enterprise",
        "cloud":             "Software",
        "new dispatch p1":   "Dispatch P1",
        "new dispatch p2":   "Dispatch P2",
        "handover in":       "Handover",
        "handover-in":       "Handover",
    }

    def _fix(val):
        raw = str(val).replace("\xa0", "").strip()
        if not raw or raw.lower() in ("nan", "none", "nat"):
            return ""
        norm  = " ".join(raw.lower().split())
        alias = _ALIASES.get(norm)
        if alias:
            return alias
        if norm.startswith("new ") and "dispatch" in norm:
            return norm[4:].strip().title()
        titled = raw.title()
        return _ACRONYMS.get(titled, titled)

    frames = []
    for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
        if sheet_name not in wb.sheetnames:
            continue
        ws   = wb[sheet_name]
        rows = list(ws.values)
        if len(rows) < 2:
            continue
        header = rows[0]
        data   = rows[1:]
        seen = {}
        clean_header = []
        for col in header:
            col_str = str(col) if col is not None else "unnamed"
            if col_str in seen:
                seen[col_str] += 1
                clean_header.append(f"{col_str}_{seen[col_str]}")
            else:
                seen[col_str] = 0
                clean_header.append(col_str)
        df = pd.DataFrame(data, columns=clean_header).reset_index(drop=True)
        # Normalise categories
        for col in ("Vertical", "Technology", "Case Delivery Type"):
            if col in df.columns:
                df[col] = df[col].astype(str).apply(_fix)
                df[col] = df[col].replace("", pd.NA)
        frames.append(df)

    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=COLUMNS)

    if "Charts" in wb.sheetnames:
        del wb["Charts"]
    wc = wb.create_sheet("Charts")
    wc.column_dimensions["A"].width = 26
    wc.column_dimensions["B"].width = 12

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_block(col_name, title, start_row):
        if col_name not in combined.columns or combined.empty:
            return start_row
        counts = (
            combined[col_name]
            .dropna()
            .astype(str)
            .str.strip()
            .pipe(lambda s: s[s != ""])
            .value_counts()
            .sort_index()
        )
        if counts.empty:
            return start_row
        t = wc.cell(row=start_row, column=1, value=title)
        t.font = Font(bold=True, size=13, color="1F3864", name="Arial")
        start_row += 1
        for c, val in [(1, col_name), (2, "Count")]:
            cell = wc.cell(row=start_row, column=c, value=val)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill      = PatternFill("solid", start_color="4472C4")
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
        for i, (label, count) in enumerate(counts.items(), 1):
            r  = start_row + i
            lc = wc.cell(row=r, column=1, value=label)
            cc = wc.cell(row=r, column=2, value=int(count))
            for cell in [lc, cc]:
                cell.font      = Font(name="Arial", size=10)
                cell.border    = border
                cell.alignment = Alignment(
                    horizontal="left" if cell.column == 1 else "center"
                )
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="DCE6F1")
        data_end = start_row + len(counts)
        chart = BarChart()
        chart.type     = "col"
        chart.grouping = "clustered"
        chart.title    = title
        chart.y_axis.title = "Cases"
        chart.width    = 18
        chart.height   = 12
        chart.style    = 10
        chart.add_data(
            Reference(wc, min_col=2, max_col=2, min_row=start_row, max_row=data_end),
            titles_from_data=True
        )
        chart.set_categories(
            Reference(wc, min_col=1, min_row=start_row + 1, max_row=data_end)
        )
        chart.series[0].graphicalProperties.solidFill = CHART_COLORS[0]
        wc.add_chart(chart, f"D{start_row}")
        return data_end + 24

    row = 1
    row = write_block("Vertical",           "Cases by Vertical",           row)
    row = write_block("Technology",         "Cases by Technology",         row)
    row = write_block("Case Delivery Type", "Cases by Case Delivery Type", row)


def _find_last_data_row(ws):
    """
    Find the true last row that contains data.
    openpyxl's ws.max_row can be inflated by formatting/ghost rows.
    We scan from the bottom up for the first non-empty row.
    """
    for row in range(ws.max_row, 0, -1):
        for cell in ws[row]:
            if cell.value is not None and str(cell.value).strip():
                return row
    return 1  # only header row exists


def _row_has_missing_fields(ws, row_idx):
    return any(
        not str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
        for col_idx in CRITICAL_COLS
    )


def _style_problem_row(ws, row_idx, reason):
    if reason == "missing":
        fill = MISSING_FILL
        font = MISSING_FONT
    else:
        fill = DUPLICATE_FILL
        font = DUPLICATE_FONT
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font


def _clear_problem_style(ws, row_idx):
    normal_font = Font(name="Arial", size=10)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill_color = getattr(cell.fill.fgColor, "rgb", None)
        if fill_color in PROBLEM_FILL_COLORS:
            cell.fill = PatternFill(fill_type=None)
            cell.font = normal_font


def _date_key(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    except Exception:
        try:
            return pd.to_datetime(raw, dayfirst=True).strftime("%Y-%m-%d")
        except Exception:
            return " ".join(raw.lower().split())


def _case_row_locations(wb, case_number, case_date):
    if not case_number:
        return []
    case_number = str(case_number).strip()
    case_date_key = _date_key(case_date)
    matches = []
    for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        last_row = _find_last_data_row(ws)
        for row_idx in range(2, last_row + 1):
            existing_date = ws.cell(row=row_idx, column=1).value
            existing_case = str(ws.cell(row=row_idx, column=2).value or "").strip()
            if existing_case == case_number and _date_key(existing_date) == case_date_key:
                matches.append((sheet_name, row_idx))
    return matches


def _highlight_problem_rows(wb):
    """
    Re-apply problem highlighting across data sheets.
    Red = missing required extraction field.
    Amber = duplicate Case# on the same date.
    """
    case_date_locations = {}
    for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        last_row = _find_last_data_row(ws)
        for row_idx in range(2, last_row + 1):
            case_number = str(ws.cell(row=row_idx, column=2).value or "").strip()
            date_key = _date_key(ws.cell(row=row_idx, column=1).value)
            if case_number and date_key:
                case_date_locations.setdefault(
                    (date_key, case_number), []
                ).append((sheet_name, row_idx))

    duplicate_locations = {
        loc
        for locations in case_date_locations.values()
        if len(locations) > 1
        for loc in locations
    }

    for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        last_row = _find_last_data_row(ws)
        for row_idx in range(2, last_row + 1):
            _clear_problem_style(ws, row_idx)
            if _row_has_missing_fields(ws, row_idx):
                _style_problem_row(ws, row_idx, "missing")
            elif (sheet_name, row_idx) in duplicate_locations:
                _style_problem_row(ws, row_idx, "duplicate")


def _check_and_archive_if_needed():
    """
    If the current tracker file spans more than 3 calendar months,
    archive it to output/archive/ and start a fresh file.
    Called at the start of every append_to_excel.
    """
    if not OUTPUT_FILE.exists():
        return

    try:
        wb = load_workbook(OUTPUT_FILE, read_only=True)
        all_dates = []
        for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
            if sheet_name not in wb.sheetnames:
                continue
            ws   = wb[sheet_name]
            rows = list(ws.values)
            if len(rows) < 2:
                continue
            header = list(rows[0])
            if "Date" not in header:
                continue
            date_idx = header.index("Date")
            for row in rows[1:]:
                val = row[date_idx]
                if val is not None:
                    try:
                        all_dates.append(pd.Timestamp(val))
                    except Exception:
                        pass
        try:
            wb.close()
        except Exception:
            pass

        if len(all_dates) < 2:
            return

        valid = [d for d in all_dates if pd.notna(d)]
        if not valid:
            return

        min_date = min(valid)
        max_date = max(valid)

        # Count distinct calendar months spanned
        months_spanned = (
            (max_date.year - min_date.year) * 12
            + (max_date.month - min_date.month)
            + 1
        )

        if months_spanned <= 3:
            return  # still within the 3-month window

        # Archive the current file
        archive_dir = OUTPUT_FILE.parent / "archive"
        archive_dir.mkdir(exist_ok=True)
        start_str = min_date.strftime("%b%Y")
        end_str   = max_date.strftime("%b%Y")
        archive_name = f"Weekend_Cases_Tracker_{start_str}_{end_str}.xlsx"
        archive_path = archive_dir / archive_name

        import shutil
        shutil.move(str(OUTPUT_FILE), str(archive_path))
        print(f"[excel] 3-month window complete — archived to archive/{archive_name}")
        print(f"[excel] Starting fresh tracker file.")

    except Exception as e:
        print(f"[excel] Archive check failed (non-fatal): {e}")


def append_to_excel(data, dt=None):
    # Check if 3-month window is complete and archive if needed
    _check_and_archive_if_needed()

    sheet_name = get_sheet_name(dt)
    wb = _ensure_workbook()
    ws = wb[sheet_name]

    new_row = [
        data.get("Date", ""),
        data.get("Case#", ""),
        data.get("Customer", ""),
        data.get("Vertical", ""),
        data.get("Technology", ""),
        data.get("Case Delivery Type", ""),
        data.get("EM", ""),
        data.get("Comments", ""),
    ]

    case_number = str(new_row[1] or "").strip()
    case_date = new_row[0]
    duplicate_rows = _case_row_locations(wb, case_number, case_date)
    if duplicate_rows:
        _highlight_problem_rows(wb)
        for dup_sheet, dup_row in duplicate_rows:
            if not _row_has_missing_fields(wb[dup_sheet], dup_row):
                _style_problem_row(wb[dup_sheet], dup_row, "duplicate")
        wb.save(OUTPUT_FILE)
        locations = ", ".join(
            f"{dup_sheet} row {dup_row}" for dup_sheet, dup_row in duplicate_rows
        )
        print(
            f"       [excel] Duplicate Case# {case_number} on {case_date} found at "
            f"{locations} — new row not saved"
        )
        return False

    # Write to the row immediately after the last real data row
    # This avoids the gap caused by openpyxl max_row inflation
    last_row    = _find_last_data_row(ws)
    target_row  = last_row + 1
    for col_idx, value in enumerate(new_row, start=1):
        ws.cell(row=target_row, column=col_idx, value=value)

    # Apply border to every cell in the new row
    thin_side = Side(style="thin", color="BFBFBF")
    row_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side,  bottom=thin_side
    )
    normal_font = Font(name="Arial", size=10)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.border = row_border
        if not cell.font or cell.font.name == "Calibri":
            cell.font = normal_font

    has_blank = _row_has_missing_fields(ws, target_row)
    if has_blank:
        _style_problem_row(ws, target_row, "missing")
        print(f"       [excel] Row {target_row} flagged red (blank fields detected)")

    _highlight_problem_rows(wb)
    _adjust_column_width(ws)
    # Save row first — nothing below can block this
    wb.save(OUTPUT_FILE)
    print(f"       [excel] Row saved to sheet '{sheet_name}' row {target_row} ✓")
    # Best-effort chart rebuild
    try:
        wb2 = load_workbook(OUTPUT_FILE)
        _rebuild_charts_sheet(wb2)
        _highlight_problem_rows(wb2)
        wb2.save(OUTPUT_FILE)
    except Exception as e:
        print(f"       [excel] Chart rebuild skipped: {e}")
    return True


def technology_missing_for_case(case_number):
    if not case_number or not OUTPUT_FILE.exists():
        return False
    try:
        wb = load_workbook(OUTPUT_FILE, read_only=True)
    except Exception:
        return False
    for sheet_name in (SHEET_SATURDAY, SHEET_SUNDAY):
        if sheet_name not in wb.sheetnames:
            continue
        ws     = wb[sheet_name]
        rows   = list(ws.values)
        if len(rows) < 2:
            continue
        header = list(rows[0])
        try:
            case_idx = header.index("Case#")
            tech_idx = header.index("Technology")
        except ValueError:
            continue
        for row in reversed(rows[1:]):
            if str(row[case_idx]).strip() == str(case_number).strip():
                tech = row[tech_idx]
                try:
                    wb.close()
                except Exception:
                    pass
                return tech is None or not str(tech).strip()
    try:
        wb.close()
    except Exception:
        pass
    return False
